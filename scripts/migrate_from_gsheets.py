"""
migrate_from_gsheets.py

One-time migration script to import historical data from the exported
Google Sheets Excel file into the new PostgreSQL/SQLite database.

Usage:
    python -m scripts.migrate_from_gsheets path/to/Microbiome_Optimization__Weekly_Plan.xlsx

The script:
1. Reads all "Plants - {date}" sheets
2. Creates Week records for each sheet
3. Creates Entry records for each veggie per person
4. Skips the TEMPLATE sheet and any empty columns
"""
import sys
import re
import datetime
import asyncio
import openpyxl
from pathlib import Path

# Add parent to path so we can import the app
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from backend.app.database import engine, Base, async_session
from backend.app.models import User, Week, Entry
from backend.app.services.item_service import normalize_item
from backend.app.services.auth_service import hash_pin

# Participant column mapping from the original spreadsheet
PARTICIPANT_COLS = {
    "Julie": 2,  # Column B
    "Wika": 3,   # Column C
    "Mike": 4,   # Column D
}

DEFAULT_PIN = "1234"  # Users should change after first login


def _all_md_interpretations(digits: str) -> set[tuple[int, int]]:
    """Return all valid (month, day) pairs the digit string could represent."""
    results = set()
    def try_add(m_str, d_str):
        try:
            m, d = int(m_str), int(d_str)
            if 1 <= m <= 12 and 1 <= d <= 31:
                results.add((m, d))
        except ValueError:
            pass

    if len(digits) <= 2:
        try_add(digits[0], digits[1:] if len(digits) > 1 else "1")
    elif len(digits) == 3:
        try_add(digits[0], digits[1:])   # M/DD
        try_add(digits[:2], digits[2:])  # MM/D
    elif len(digits) == 4:
        try_add(digits[:2], digits[2:])  # MM/DD
    return results


def _resolve_sheet_date(sheet_name: str, ws) -> datetime.date | None:
    """
    Return the best date for a sheet, preferring B1 when it's consistent with
    the sheet name's digits (B1 has the correct year; sheet name does not).
    Falls back to a year-guessing heuristic when B1 is absent or corrupt.
    """
    name_match = re.match(r"Plants\s*-\s*(\d+)", sheet_name)
    if not name_match:
        return None
    digits = name_match.group(1)

    # Try B1 first — it contains the year when valid.
    b1_value = ws.cell(row=1, column=2).value
    if isinstance(b1_value, datetime.datetime):
        b1_date = b1_value.date()
    elif isinstance(b1_value, datetime.date):
        b1_date = b1_value
    else:
        b1_date = None

    if b1_date and (b1_date.month, b1_date.day) in _all_md_interpretations(digits):
        return b1_date  # B1 is consistent with the sheet name — trust it.

    # B1 is absent or corrupt (e.g. wrong date copied from another sheet).
    # Fall back to the sheet name digits with a year-guessing heuristic:
    # pick the most-recent past date to avoid placing old sheets in future years.
    interpretations = _all_md_interpretations(digits)
    if not interpretations:
        return None

    today = datetime.date.today()
    best_past: datetime.date | None = None
    best_future: datetime.date | None = None

    for month, day in interpretations:
        for year in [today.year, today.year - 1, today.year - 2]:
            try:
                d = datetime.date(year, month, day)
                if d <= today:
                    if best_past is None or d > best_past:
                        best_past = d
                    break
                elif best_future is None or d < best_future:
                    best_future = d
            except ValueError:
                continue

    return best_past if best_past is not None else best_future


async def migrate(excel_path: str):
    """Run the migration."""
    print(f"Loading workbook from {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Create tables
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    
    async with async_session() as db:
        # Create users if they don't exist
        users = {}
        for name in PARTICIPANT_COLS:
            from sqlalchemy import select
            result = await db.execute(select(User).where(User.name == name))
            user = result.scalar_one_or_none()
            if not user:
                user = User(
                    name=name,
                    pin_hash=hash_pin(DEFAULT_PIN),
                    gender="male" if name == "Mike" else "female",
                    email=None,
                )
                db.add(user)
                await db.flush()
                print(f"  Created user: {name} (id={user.id})")
            users[name] = user
        
        total_entries = 0
        weeks_created = 0
        
        for sheet_name in wb.sheetnames:
            if "TEMPLATE" in sheet_name.upper():
                continue
            
            ws = wb[sheet_name]

            week_start = _resolve_sheet_date(sheet_name, ws)
            if not week_start:
                print(f"  Skipping sheet '{sheet_name}' — couldn't parse date")
                continue
            
            # Ensure it's a Sunday (adjust if needed)
            days_since_sunday = (week_start.weekday() + 1) % 7
            if days_since_sunday != 0:
                week_start = week_start - datetime.timedelta(days=days_since_sunday)
            week_end = week_start + datetime.timedelta(days=6)
            
            # Get or create week
            from sqlalchemy import select as sa_select
            week_result = await db.execute(sa_select(Week).where(Week.start_date == week_start))
            week = week_result.scalar_one_or_none()
            if not week:
                week = Week(start_date=week_start, end_date=week_end, is_active=False)
                db.add(week)
                await db.flush()
                weeks_created += 1
            
            # Import entries for each participant
            for name, col_idx in PARTICIPANT_COLS.items():
                # Verify name in row 2
                cell_name = ws.cell(row=2, column=col_idx).value
                if not cell_name or name.lower() not in str(cell_name).lower():
                    continue
                
                # Read veggies from row 4 downward
                for row in range(4, ws.max_row + 1):
                    val = ws.cell(row=row, column=col_idx).value
                    if not val or not str(val).strip():
                        continue
                    
                    item_name = str(val).strip()
                    normalized = normalize_item(item_name)
                    
                    # Check for duplicate in this week
                    dup_result = await db.execute(
                        sa_select(Entry).where(
                            Entry.user_id == users[name].id,
                            Entry.week_id == week.id,
                            Entry.item_name_normalized == normalized,
                        )
                    )
                    if dup_result.scalar_one_or_none():
                        continue  # Skip duplicate
                    
                    entry = Entry(
                        user_id=users[name].id,
                        week_id=week.id,
                        item_name=item_name,
                        item_name_normalized=normalized,
                    )
                    db.add(entry)
                    total_entries += 1
            
            print(f"  Imported sheet '{sheet_name}' (week of {week_start})")
        
        await db.commit()
        print(f"\nMigration complete!")
        print(f"  Weeks created: {weeks_created}")
        print(f"  Entries imported: {total_entries}")
        print(f"  Users: {', '.join(users.keys())}")
        print(f"\nDefault PIN for all users: {DEFAULT_PIN}")
        print(f"Please change PINs after first login!")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python -m scripts.migrate_from_gsheets <path_to_excel>")
        sys.exit(1)
    
    asyncio.run(migrate(sys.argv[1]))
